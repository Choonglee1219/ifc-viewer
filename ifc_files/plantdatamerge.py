# -*- coding: utf-8 -*-
"""
Created on Wed Jun 21 10:24:30 2023

@author: 39076_Choonghyun-Lee
"""

#%% Library import
import numpy as np
import pandas as pd
import openpyxl as op
import re
import os

#%% Dataframes from Exported ifc properties and EDB plant data
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
file1 = r"aveva_230601-properties.xlsx"
file2 = r"plant_data_sample.xlsx"
file3 = r"temp.xlsx"
sheet1 = "elementassembly"
sheet2 = "line index"

df1 = pd.read_excel(io=file1, sheet_name=sheet1)
df2 = pd.read_excel(io=file2, sheet_name=sheet2)

#%% Making Boolean Series from ifc class Name and 'tag-no' pattern
pattern = re.compile('^[0-9]'+'-')
sr1 = df1.Name.str.split('/').str[0]
cond1 = sr1.str.contains(pattern)

#%% Masking, Merge, Export to excel
df1["tag_no"] = sr1.mask(~cond1, np.nan)
df3 = df1.merge(df2, how='left', on='tag_no')
df3.to_excel(file3, index=False)

#%% Replace original sheet in properties Workbook with new sheet in new merged Workbook
wb1 = op.load_workbook(file1)
ws1 = wb1["elementassembly"]

wb3 = op.load_workbook(file3)
ws3 = wb3["Sheet1"]

row_no = 1
for row in ws3.iter_rows():
    col_no = 1
    for cell in row:
        ws1.cell(row_no, col_no).value = cell.value
        col_no += 1
    row_no += 1

wb1.save(file1)

if os.path.exists(file3):
    os.remove(file3)
else:
    pass